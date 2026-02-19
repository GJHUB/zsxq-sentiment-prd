"""报告生成模块 - Excel报告"""

import logging
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.filters import AutoFilter

from .config import get_config

logger = logging.getLogger(__name__)

OUTLOOK_COLORS = {
    "看多": "C6EFCE",
    "看空": "FFC7CE",
    "中性": "FFEB9C",
    "分歧": "B4C6E7",
}

HEADERS = [
    "时间", "作者", "群主帖子", "是否财经", "金融产品", "具体标的",
    "看法", "群主看法", "群主理由", "原因分析", "核心观点", "评论数", "帖子摘要",
]


class ReportGenerator:
    """Excel报告生成器"""

    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or get_config("output_dir", "output")
        os.makedirs(self.output_dir, exist_ok=True)

    def generate(self, analysis, topics, date=None, group_names=None):
        date = date or datetime.now().strftime("%Y-%m-%d")
        filename = f"舆情分析_{date}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        group_names = group_names or {}

        try:
            wb = Workbook()
            wb.remove(wb.active)

            if "group_id" in analysis.columns:
                group_ids = analysis["group_id"].unique().tolist()
            else:
                group_ids = ["default"]
                analysis["group_id"] = "default"

            for gid in group_ids:
                name = group_names.get(gid, str(gid))
                short_name = name[:14] if len(name) > 14 else name
                group_df = analysis[analysis["group_id"] == gid]
                self._create_sheet(wb, group_df, short_name, gid=gid, group_name=name)

            self._apply_styles(wb)
            wb.save(filepath)
            logger.info("报告已生成: %s", filepath)
            return filepath
        except Exception as e:
            logger.error("Excel生成失败，降级到CSV: %s", e)
            return self._fallback_csv(analysis, date)

    def _create_sheet(self, wb, df, title, gid="", group_name=""):
        ws = wb.create_sheet(title[:31])

        # 第一行：星球信息
        ws.append([f"星球: {group_name}  (ID: {gid})" if group_name else f"ID: {gid}"])
        # 第二行：字段名
        ws.append(HEADERS)

        if df.empty:
            ws.append(["暂无数据"])
            return

        for _, row in df.iterrows():
            targets_str = self._targets_to_str(row.get("targets", []))
            is_financial = "是" if row.get("is_financial") else "否"
            ws.append([
                str(row.get("create_time", ""))[:19],
                row.get("author", ""),
                "是" if row.get("is_owner_post") else "否",
                is_financial,
                row.get("product_type", ""),
                targets_str,
                row.get("outlook", ""),
                row.get("owner_outlook", "无"),
                row.get("owner_reason", "无"),
                row.get("reason", ""),
                row.get("summary", ""),
                row.get("comments_count", 0),
                row.get("post_excerpt", "")[:200],
            ])

        # 添加自动筛选（基于第二行表头，范围到最后一行）
        last_col_letter = chr(ord("A") + len(HEADERS) - 1)
        ws.auto_filter.ref = f"A2:{last_col_letter}{ws.max_row}"
        # 默认筛选：过滤掉非财经（只显示"是"）
        # "是否财经"是第4列（index 3）
        ws.auto_filter.add_filter_column(3, ["是"])

    @staticmethod
    def _targets_to_str(targets) -> str:
        if isinstance(targets, list):
            return "、".join(str(t) for t in targets)
        return str(targets) if targets else ""

    def _apply_styles(self, wb):
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        info_font = Font(bold=True, size=12, color="333333")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for ws in wb.worksheets:
            # 第一行：星球信息
            for cell in ws[1]:
                cell.font = info_font
                cell.alignment = Alignment(vertical="center")

            # 第二行：表头
            if ws.max_row >= 2:
                for cell in ws[2]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

            # 数据行
            for row_cells in ws.iter_rows(min_row=3, max_row=ws.max_row):
                for cell in row_cells:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            # 自动列宽
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value or "")
                        length = sum(2 if ord(c) > 127 else 1 for c in val)
                        max_length = max(max_length, length)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

            self._apply_outlook_colors(ws)

    def _apply_outlook_colors(self, ws):
        if ws.max_row < 2:
            return
        header_row = [cell.value for cell in ws[2]]
        for idx, h in enumerate(header_row):
            if h == "看法":
                for row_cells in ws.iter_rows(min_row=3, max_row=ws.max_row):
                    cell = row_cells[idx]
                    color = OUTLOOK_COLORS.get(str(cell.value or ""))
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _fallback_csv(self, df, date):
        filename = f"舆情分析_{date}.csv"
        filepath = os.path.join(self.output_dir, filename)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")
        logger.info("CSV报告已生成（降级）: %s", filepath)
        return filepath
